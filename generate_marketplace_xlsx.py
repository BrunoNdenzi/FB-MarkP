"""
Facebook Marketplace Bulk Upload XLSX Generator
Project: Tile Warehouse Sales — Charlotte NC

MATCHES the exact Facebook Marketplace Bulk Upload Template format:
  Row 1 : Header title
  Row 2 : Note
  Row 3 : Column requirement descriptions
  Row 4 : Column headers  (TITLE | PRICE | CONDITION | DESCRIPTION | CATEGORY)
  Row 5+: Listing data rows

HOW TO USE:
  1. pip install openpyxl
  2. python generate_marketplace_xlsx.py
  3. Output: marketplace_bulk_upload.xlsx
  4. Upload via:
       facebook.com/marketplace > Sell > Create multiple listings > Upload file

HOW TO UPDATE:
  - Change any price, title, or description in the LISTINGS section below
  - Re-run script — new file is generated instantly
  - Re-upload to Facebook Marketplace
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_FILE = "marketplace_bulk_upload.xlsx"

# ─────────────────────────────────────────────────────────
# LISTINGS — edit titles, prices, descriptions here
# PRICE must be a whole number (no decimals)
# CONDITION must be exactly: New / Used - Like New / Used - Good / Used - Fair
# TITLE must be 150 characters or fewer
# DESCRIPTION must be 5000 characters or fewer
# ──────���──────────────────────────────────────────────────
LISTINGS = [

    # ── LISTING 1 — White Marble Polished 24x48 ───────────
    {
        "title"      : "24x48 Polished Porcelain Tile - White Marble Look - $99/case - Charlotte",
        "price"      : 99,                        # UPDATE PRICE HERE
        "condition"  : "New",
        "description": (
            "Selling premium quality 24x48 polished porcelain tiles directly from warehouse. "
            "White marble look with natural grey veining. Works for floors, walls, bathrooms, "
            "kitchens, and feature walls. Brand new in boxes. Priced well below retail. "
            "$99 per case - multiple cases available. Pickup in Charlotte. Cash or Zelle. "
            "Size: 24x48 inches. Finish: Polished. Style: White marble with grey veining. "
            "Text for details and available quantity. Can hold with deposit."
        ),
        "category"   : "Home & Garden//Home Improvement",
    },

    # ── LISTING 2 — Calacatta Gold Polished 24x48 ─────────
    {
        "title"      : "24x48 Polished Porcelain Tile - Calacatta Gold Look - $99/case - Charlotte",
        "price"      : 99,                        # UPDATE PRICE HERE
        "condition"  : "New",
        "description": (
            "Warehouse surplus sale. 24x48 polished porcelain tiles with warm white base and "
            "gold and brown veining. Upscale look for kitchens, master bathrooms, or living areas. "
            "Brand new in boxes. Limited quantity. $99 per case. Pickup Charlotte area. "
            "Cash or Zelle. Size: 24x48 inches. Finish: Polished. Style: Calacatta Gold veining. "
            "Text to confirm availability."
        ),
        "category"   : "Home & Garden//Home Improvement",
    },

    # ── LISTING 3 — Dark Charcoal Matte 24x48 ─────────────
    {
        "title"      : "24x48 Dark Charcoal Porcelain Tile - Matte Finish - $70/case - Charlotte",
        "price"      : 70,                        # UPDATE PRICE HERE
        "condition"  : "New",
        "description": (
            "Warehouse clearance. Large format 24x48 dark charcoal slate look porcelain tile, "
            "matte finish. Perfect for modern bathrooms, commercial spaces, accent walls, or outdoor areas. "
            "Brand new in boxes. $70 per case. Pickup Charlotte area. Cash or Zelle. "
            "Size: 24x48 inches. Finish: Matte. Style: Dark charcoal slate look. "
            "Text for availability and quantity."
        ),
        "category"   : "Home & Garden//Home Improvement",
    },

    # ── LISTING 4 — Wood-Look Plank ────────────────────────
    # PLACEHOLDER PRICE — update once warehouse confirms
    {
        "title"      : "Wood-Look Plank Flooring - Warehouse Clearance Charlotte - Message for Price",
        "price"      : 50,                        # PLACEHOLDER — update once confirmed
        "condition"  : "New",
        "description": (
            "Warehouse surplus. Wood-look plank tile or luxury vinyl plank flooring. "
            "Warm brown tones, realistic wood grain look. Brand new in boxes. Multiple cases available. "
            "Price varies by quantity - message us with how many cases you need and we will confirm current pricing. "
            "Pickup in Charlotte. Cash or Zelle. Text for pricing and availability."
        ),
        "category"   : "Home & Garden//Home Improvement",
    },

    # ── LISTING 5 — Blue-Grey Large Format ────────────────
    # PLACEHOLDER PRICE — update once warehouse confirms
    {
        "title"      : "Large Format Blue-Grey Porcelain Tile - Clearance Price - Charlotte",
        "price"      : 50,                        # PLACEHOLDER — update once confirmed
        "condition"  : "New",
        "description": (
            "Warehouse clearance sale. Large format blue-grey porcelain tile. "
            "Cool modern tone, great for bathrooms, kitchens, or accent walls. "
            "Brand new in boxes. Price varies - message us with quantity needed and we will confirm current pricing. "
            "Pickup Charlotte area. Cash or Zelle. Text for pricing and availability."
        ),
        "category"   : "Home & Garden//Home Improvement",
    },

    # ── LISTING 6 — Catch-All / Traffic Driver ─────────────
    {
        "title"      : "Charlotte Tile Warehouse Sale - Multiple Styles - 24x48 and 12x24 Available",
        "price"      : 12,
        "condition"  : "New",
        "description": (
            "Warehouse surplus tile inventory in Charlotte. Multiple styles available including "
            "white marble look, dark charcoal, wood-look flooring, and more. "
            "24x48 Polished $99/case. 24x48 Matte $70/case. 12x24 Polished $22/case. 12x24 Matte $12/case. "
            "Other styles - message for pricing. New in boxes. Priced below retail. "
            "Great for contractors, flippers, and homeowners. Pickup Charlotte area. Cash or Zelle. "
            "Text with what you are looking for and we will confirm availability."
        ),
        "category"   : "Home & Garden//Home Improvement",
    },

]


# ─────────────────────────────────────────────────────────
# XLSX ENGINE — do not edit below this line
# ─────────────────────────────────────────────────────────
def generate_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Upload Template"

    # Row 1 — Template title (bold)
    ws["A1"] = "Facebook Marketplace Bulk Upload Template"
    ws["A1"].font = Font(bold=True, size=13)

    # Row 2 — Instruction note (red italic, matches Facebook template)
    ws["A2"] = (
        "You can create up to 50 listings at once. "
        "When you are finished, be sure to save or export this as an XLS/XLSX file."
    )
    ws["A2"].font = Font(italic=True, color="FF0000")

    # Row 3 — Column requirement hints (grey small text)
    hints = [
        "REQUIRED | Plain text (up to 150 characters)",
        "REQUIRED | A whole number in $",
        'REQUIRED | Supported values: "New"; "Used - Like New"; "Used - Good"; "Used - Fair"',
        "OPTIONAL | Plain text (up to 5000 characters)",
        "OPTIONAL | Type of listing",
    ]
    for col, hint in enumerate(hints, start=1):
        cell = ws.cell(row=3, column=col, value=hint)
        cell.font = Font(color="808080", size=9)

    # Row 4 — Column headers (bold)
    headers = ["TITLE", "PRICE", "CONDITION", "DESCRIPTION", "CATEGORY"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)

    # Rows 5+ — Listing data
    for i, listing in enumerate(LISTINGS, start=5):
        ws.cell(row=i, column=1, value=listing["title"])
        ws.cell(row=i, column=2, value=listing["price"])
        ws.cell(row=i, column=3, value=listing["condition"])
        ws.cell(row=i, column=4, value=listing["description"])
        ws.cell(row=i, column=5, value=listing["category"])
        # Wrap text in description column
        ws.cell(row=i, column=4).alignment = Alignment(wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 35

    wb.save(OUTPUT_FILE)
    print(f"\nGenerated : {OUTPUT_FILE}")
    print(f"Listings  : {len(LISTINGS)}")
    print(f"\nNext step : Upload via Facebook Marketplace")
    print(f"  facebook.com/marketplace > Sell > Create multiple listings > Upload file\n")


if __name__ == "__main__":
    generate_xlsx()
